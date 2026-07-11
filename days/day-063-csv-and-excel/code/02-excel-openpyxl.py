#!/usr/bin/env python3
"""
Day 063 - CSV 与 Excel
示例 2: openpyxl Excel 读写与样式

本示例演示 openpyxl 的核心功能：
1. 创建工作簿和工作表
2. 写入数据与格式化样式
3. 单元格合并、条件格式
4. 图表创建（柱状图等）
5. 读取现有 Excel 文件
6. 大文件优化技巧

安装依赖: pip install openpyxl
运行方式: python3 02-excel-openpyxl.py
"""

from pathlib import Path
from datetime import datetime

# ════════════════════════════════════════════
# 1. 基础操作：创建 Excel 文件
# ════════════════════════════════════════════

print("=" * 60)
print("📊 openpyxl — Excel 文件读写")
print("=" * 60)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    HAS_OPENPYXL = True
except ImportError as e:
    HAS_OPENPYXL = False
    print(f"⚠️  openpyxl 未安装: {e}")
    print("   安装: pip install openpyxl")

if HAS_OPENPYXL:
    # ─── 创建工作簿 ───

    print("\n--- 1. 创建基础 Excel 工作簿 ---")

    wb = Workbook()
    ws = wb.active
    ws.title = "销售报表"

    # 添加第二个工作表
    ws2 = wb.create_sheet("数据明细")

    print(f"工作表列表: {wb.sheetnames}")

    # ─── 2. 写入数据与样式 ───

    print("\n--- 2. 写入数据与样式 ---")

    # 表头数据
    headers = ["序号", "产品名称", "分类", "销量", "单价(元)", "总金额", "日期"]
    ws.append(headers)

    # 销售数据
    sales_data = [
        [1, "iPhone 16 Pro", "手机", 120, 8999, 1079880, datetime.now()],
        [2, "MacBook Air M4", "笔记本", 45, 9499, 427455, datetime.now()],
        [3, "iPad Air", "平板", 68, 4799, 326332, datetime.now()],
        [4, "AirPods Pro 3", "耳机", 200, 1999, 399800, datetime.now()],
        [5, "Apple Watch S10", "手表", 85, 3199, 271915, datetime.now()],
        [6, "iMac 27\"", "台式机", 12, 14999, 179988, datetime.now()],
        [7, "Mac mini M4", "迷你主机", 30, 4799, 143970, datetime.now()],
        [8, "Vision Pro", "头显", 8, 29999, 239992, datetime.now()],
    ]

    for row_data in sales_data:
        ws.append(row_data)

    print(f"已写入 {len(sales_data)} 条销售记录")

    # ─── 3. 样式美化 ───

    print("\n--- 3. 样式美化 ---")

    # 定义样式
    header_font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    center_alignment = Alignment(horizontal="center", vertical="center")

    # 应用表头样式
    for cell in ws[1]:  # 第 1 行（表头）
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 交替行颜色
    light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_alignment

            # 偶数行浅色背景
            if row_idx % 2 == 0:
                cell.fill = light_fill

    # 金额列（F列）格式：￥#,##0.00
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=6).number_format = '¥#,##0.00'
        ws.cell(row=row, column=5).number_format = '¥#,##0'

    # 日期列格式
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).number_format = 'YYYY-MM-DD HH:MM'

    print("样式应用完成")

    # ─── 4. 列宽自动调整 ───

    print("\n--- 4. 列宽自动调整 ---")

    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                val = str(cell.value or "")
                # 中文字符算 2 个字符宽度
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > max_length:
                    max_length = length
            except:
                pass

        adjusted_width = min(max_length + 4, 40)  # 最大 40
        ws.column_dimensions[column_letter].width = adjusted_width

    print("列宽已自动调整")

    # ─── 5. 合并单元格（标题行） ───

    print("\n--- 5. 合并单元格 ---")

    # 在第一行前插入一个标题行
    ws.insert_rows(1)
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"产品销售统计报表 ({datetime.now().strftime('%Y-%m-%d')})"
    title_cell.font = Font(name="Microsoft YaHei", bold=True,
                            size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    print("标题行已添加")

    # ─── 6. 创建图表 ───

    print("\n--- 6. 创建图表 ---")

    # 柱状图：产品销量对比
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "产品销售数量对比"
    bar_chart.y_axis.title = "销量 (台)"
    bar_chart.x_axis.title = "产品名称"
    bar_chart.style = 10

    # 数据范围（销量列 D，跳过标题行）
    data_ref = Reference(ws, min_col=4, min_row=3, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=2, min_row=3, max_row=ws.max_row)

    bar_chart.add_data(data_ref, titles_from_data=False)
    bar_chart.set_categories(cats_ref)
    bar_chart.shape = 4

    # 设置图表大小
    bar_chart.width = 25
    bar_chart.height = 15

    # 添加图表到工作表
    ws.add_chart(bar_chart, "A12")

    print("柱状图已添加")

    # 饼图：各分类销售额占比
    print("\n--- 创建饼图 ---")

    # 按分类汇总（简化：用产品级数据）
    pie_chart = PieChart()
    pie_chart.title = "产品销售额占比"
    pie_chart.style = 10

    pie_data = Reference(ws, min_col=6, min_row=3, max_row=ws.max_row)
    pie_cats = Reference(ws, min_col=2, min_row=3, max_row=ws.max_row)

    pie_chart.add_data(pie_data, titles_from_data=False)
    pie_chart.set_categories(pie_cats)

    # 添加数据标签
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.dataLabels.showCatName = True
    pie_chart.dataLabels.showVal = False

    pie_chart.width = 20
    pie_chart.height = 15

    ws.add_chart(pie_chart, "A30")

    print("饼图已添加")

    # ─── 7. 保存 ───

    print("\n--- 7. 保存文件 ---")

    output_path = Path("/tmp/day063_excel_report.xlsx")
    wb.save(str(output_path))
    print(f"Excel 文件已保存: {output_path}")
    print(f"文件大小: {output_path.stat().st_size / 1024:.1f} KB")

    # ─── 8. 读取 Excel ───

    print("\n\n--- 8. 读取 Excel 文件 ---")

    loaded_wb = load_workbook(str(output_path), data_only=True)
    loaded_ws = loaded_wb.active

    print(f"工作表: {loaded_ws.title}")
    print(f"数据范围: {loaded_ws.dimensions}")
    print(f"行数: {loaded_ws.max_row}, 列数: {loaded_ws.max_column}")

    print("\n前 3 行数据:")
    for row in loaded_ws.iter_rows(min_row=1, max_row=3, values_only=True):
        print(f"  {row}")

    # 按行读取
    print("\n产品列表:")
    for row in loaded_ws.iter_rows(min_row=3, min_col=2,
                                     max_col=4, values_only=True):
        if row[0]:
            print(f"  {row[0]:20s} | 分类: {row[2]:6s} | 销量: {row[1]:>5d}")

    loaded_wb.close()

    # ─── 9. 条件格式 ───

    print("\n\n--- 9. 条件格式 ---")

    cond_wb = Workbook()
    cond_ws = cond_wb.active
    cond_ws.title = "条件格式"

    # 写入销售数据
    cond_ws.append(["销售员", "销售额", "目标额", "完成率"])
    salespeople = [
        ["张三", 85000, 100000],
        ["李四", 120000, 100000],
        ["王五", 95000, 100000],
        ["赵六", 60000, 100000],
        ["孙七", 110000, 100000],
        ["周八", 45000, 100000],
    ]
    for person in salespeople:
        name, sales, target = person
        rate = sales / target
        cond_ws.append([name, sales, target, rate])

    # 设置完成率百分比格式
    for row in range(2, len(salespeople) + 2):
        cond_ws.cell(row=row, column=4).number_format = '0.00%'

    # 添加条件格式：完成率 >= 100% 绿色背景
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")

    cond_ws.conditional_formatting.add(
        "D2:D10",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   fill=green_fill, font=green_font)
    )

    # 完成率 < 80% 红色背景
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")

    cond_ws.conditional_formatting.add(
        "D2:D10",
        CellIsRule(operator="lessThan", formula=["0.8"],
                   fill=red_fill, font=red_font)
    )

    cond_path = Path("/tmp/day063_conditional.xlsx")
    cond_wb.save(str(cond_path))
    print(f"条件格式 Excel 已保存: {cond_path}")
    cond_wb.close()

    # ─── 清理 ───
    for p in [output_path, cond_path]:
        if p.exists():
            p.unlink()

print("\n✅ openpyxl 示例完成！")
