#!/usr/bin/env python3
"""
Day 063 - CSV 与 Excel
示例 3: 实战 — 报表生成器

本示例实现一个通用的报表生成器 ReportGenerator，支持：
1. 从 CSV / Excel 读取原始数据
2. 数据清洗与统计汇总
3. 生成格式化的 Excel 报表（含样式、图表）
4. 多 sheet 输出（明细表 + 汇总表 + 图表）
5. pandas 与 openpyxl 结合使用

安装依赖: pip install pandas openpyxl
运行方式: python3 03-report-generator.py
"""

from pathlib import Path
from datetime import datetime
from collections import defaultdict

# 尝试导入依赖
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import csv
    HAS_CSV = True
except ImportError:
    HAS_CSV = False

# ════════════════════════════════════════════
# ReportGenerator — 通用报表生成器
# ════════════════════════════════════════════


class ReportGenerator:
    """
    通用报表生成器
    
    功能：
    - 支持 CSV / Excel 输入
    - 自动数据清洗
    - 多维度统计汇总
    - 格式化 Excel 输出（含图表）
    - 多 sheet 布局
    """

    def __init__(self, input_path: str):
        """
        初始化报表生成器

        Args:
            input_path: 输入文件路径（.csv / .xlsx）
        """
        self.input_path = Path(input_path)
        self.data = None
        self.summary = {}
        self._load_data()

    def _load_data(self):
        """根据文件扩展名加载数据"""
        if not self.input_path.exists():
            raise FileNotFoundError(f"文件未找到: {self.input_path}")

        ext = self.input_path.suffix.lower()

        if ext == ".csv":
            self.data = self._load_csv()
        elif ext in (".xlsx", ".xls"):
            self.data = self._load_excel()
        else:
            raise ValueError(f"不支持的文件格式: {ext} (支持 .csv, .xlsx)")

    def _load_csv(self) -> list:
        """从 CSV 文件加载数据"""
        rows = []
        with open(self.input_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        return rows

    def _load_excel(self) -> list:
        """从 Excel 文件加载数据"""
        if not HAS_PANDAS:
            raise ImportError("需要安装 pandas: pip install pandas")

        df = pd.read_excel(self.input_path, engine="openpyxl")
        return df.to_dict("records")

    def clean_data(self):
        """
        数据清洗：
        - 去除空行
        - 去除前后空格
        - 数值字段转换
        - 日期字段标准化
        """
        print("🧹 开始数据清洗...")

        cleaned = []
        for row in self.data:
            # 去除空行
            if not any(row.values()):
                continue

            # 去除前后空格
            clean_row = {}
            for key, value in row.items():
                if isinstance(value, str):
                    value = value.strip()
                clean_row[key.strip()] = value

            cleaned.append(clean_row)

        print(f"  清洗前: {len(self.data)} 行, 清洗后: {len(cleaned)} 行")
        self.data = cleaned

    def analyze(self, group_by: str = None, value_fields: list = None):
        """
        数据分析与统计

        Args:
            group_by: 分组字段名
            value_fields: 数值字段列表
        """
        print(f"📊 开始数据分析...")

        if not self.data:
            print("  ⚠️  没有数据可分析")
            return

        # 基本统计
        self.summary["total_records"] = len(self.data)
        self.summary["fields"] = list(self.data[0].keys())
        self.summary["report_time"] = datetime.now()

        print(f"  总记录数: {self.summary['total_records']}")
        print(f"  字段列表: {self.summary['fields']}")

        # 分组统计
        if group_by and value_fields:
            grouped = defaultdict(lambda: {f: 0 for f in value_fields})
            counts = defaultdict(int)

            for row in self.data:
                key = row.get(group_by, "未知")
                counts[key] += 1
                for field in value_fields:
                    try:
                        grouped[key][field] += float(row.get(field, 0) or 0)
                    except (ValueError, TypeError):
                        pass

            self.summary["group_by"] = group_by
            self.summary["groups"] = []
            for key in sorted(grouped.keys()):
                group_info = {
                    "name": key,
                    "count": counts[key],
                }
                group_info.update(grouped[key])
                self.summary["groups"].append(group_info)

            print(f"  分组维度: {group_by}")
            for g in self.summary["groups"][:5]:
                vals = ", ".join(f"{k}={v:.2f}" for k, v in g.items()
                                 if k not in ("name", "count"))
                print(f"    {g['name']}: {g['count']} 条 | {vals}")

    def to_excel(self, output_path: str):
        """
        生成格式化的 Excel 报表

        包含:
        Sheet 1 - 数据明细
        Sheet 2 - 汇总统计
        Sheet 3 - 图表

        Args:
            output_path: 输出 Excel 文件路径
        """
        print(f"\n📝 生成 Excel 报表: {output_path}")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side)
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = Workbook()

        # ═══ Sheet 1: 数据明细 ═══
        ws_detail = wb.active
        ws_detail.title = "数据明细"

        if self.data:
            # 写表头
            headers = list(self.data[0].keys())
            ws_detail.append(headers)

            # 写数据
            for row in self.data:
                ws_detail.append([row.get(h, "") for h in headers])

            # 样式
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for cell in ws_detail[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 列宽
            for col in ws_detail.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws_detail.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = max_len + 4

        # ═══ Sheet 2: 汇总统计 ═══
        ws_summary = wb.create_sheet("汇总统计")

        # 标题
        ws_summary["A1"] = "报表汇总"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.merge_cells("A1:E1")

        # 基本信息
        ws_summary["A3"] = "统计信息"
        ws_summary["A3"].font = Font(bold=True, size=12)

        info_rows = [
            ("总记录数", self.summary.get("total_records", 0)),
            ("字段数", len(self.summary.get("fields", []))),
            ("维度", self.summary.get("group_by", "无")),
            ("生成时间", self.summary.get("report_time", "").strftime("%Y-%m-%d %H:%M")),
        ]

        for i, (key, val) in enumerate(info_rows):
            ws_summary.cell(row=4 + i, column=1).value = key
            ws_summary.cell(row=4 + i, column=1).font = Font(bold=True)
            ws_summary.cell(row=4 + i, column=2).value = val

        # 分组统计表
        if self.summary.get("groups"):
            start_row = 10
            ws_summary.cell(row=start_row, column=1).value = "分组统计明细"
            ws_summary.cell(row=start_row, column=1).font = Font(bold=True, size=12)

            groups = self.summary["groups"]
            # 表头
            group_headers = list(groups[0].keys())
            for j, h in enumerate(group_headers, 1):
                cell = ws_summary.cell(row=start_row + 1, column=j)
                cell.value = h
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid")

            # 数据
            for i, g in enumerate(groups):
                for j, h in enumerate(group_headers, 1):
                    val = g.get(h, "")
                    if isinstance(val, float):
                        val = round(val, 2)
                    ws_summary.cell(row=start_row + 2 + i, column=j).value = val

        # ═══ Sheet 3: 图表 ═══
        ws_chart = wb.create_sheet("图表")

        if self.summary.get("groups"):
            groups = self.summary["groups"]
            group_names = [g.get("name", "") for g in groups]
            group_headers = [k for k in groups[0].keys()
                             if k not in ("name", "count")]

            # 在 Sheet 3 写入数据用于图表
            ws_chart.append(["分组", "指标", "值"])
            for g in groups:
                for h in group_headers:
                    ws_chart.append([g.get("name", ""), h, g.get(h, 0)])

            # 柱状图: 各组各指标对比
            if group_headers:
                chart = BarChart()
                chart.type = "col"
                chart.title = f"{self.summary.get('group_by', '分组')} — 指标对比"
                chart.y_axis.title = "数值"
                chart.style = 10

                data_start = 2
                data_end = data_start + len(groups) * len(group_headers) - 1

                for idx, h in enumerate(group_headers):
                    # 每个指标一个系列
                    # 简化为使用类别
                    pass

                # 简化：创建第一个 group_headers 的对比
                first_field = group_headers[0]
                ws_chart_data = wb.create_sheet("_chart_data")

                # 写入图表源数据
                ws_chart_data.append([self.summary.get("group_by"), first_field])
                for g in groups:
                    ws_chart_data.append([g.get("name", ""), g.get(first_field, 0)])

                chart2 = BarChart()
                chart2.title = f"各{self.summary.get('group_by', '分组')} {first_field} 对比"
                chart2.y_axis.title = first_field

                data_ref = Reference(ws_chart_data, min_col=2, min_row=1,
                                     max_row=len(groups) + 1)
                cats_ref = Reference(ws_chart_data, min_col=1, min_row=2,
                                     max_row=len(groups) + 1)
                chart2.add_data(data_ref, titles_from_data=True)
                chart2.set_categories(cats_ref)
                chart2.width = 20
                chart2.height = 12

                ws_chart.add_chart(chart2, "A1")

                # 删除临时 sheet
                del wb["_chart_data"]

        # 保存
        wb.save(output_path)
        print(f"✅ 报表已生成: {output_path}")

    def to_csv(self, output_path: str):
        """导出为 CSV"""
        if not self.data:
            print("⚠️  没有数据可导出")
            return

        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=self.data[0].keys())
            writer.writeheader()
            writer.writerows(self.data)

        print(f"✅ CSV 已导出: {output_path}")


# ════════════════════════════════════════════
# 主程序演示
# ════════════════════════════════════════════

def main():
    print("=" * 60)
    print("🚀 实战：报表生成器演示")
    print("=" * 60)

    tmp_dir = Path("/tmp/day063_report")
    tmp_dir.mkdir(parents=True, exist_ok=True)

    # ─── 1. 生成模拟销售数据 ───

    print("\n--- 1. 生成模拟销售数据 ---")

    import random
    random.seed(42)

    products = [
        ("iPhone 16 Pro", "手机", 8999),
        ("MacBook Air M4", "笔记本", 9499),
        ("iPad Air", "平板", 4799),
        ("AirPods Pro 3", "耳机", 1999),
        ("Apple Watch S10", "手表", 3199),
        ("iMac 27\"", "台式机", 14999),
        ("Mac mini M4", "迷你主机", 4799),
        ("Vision Pro", "头显", 29999),
    ]

    regions = ["华北", "华东", "华南", "西南", "华中"]
    salespeople = ["张三", "李四", "王五", "赵六", "孙七", "周八"]

    csv_path = tmp_dir / "sales_data.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["订单号", "产品名", "分类", "区域", "销售员",
                         "销量", "单价", "总金额", "日期"])

        for i in range(500):
            product = random.choice(products)
            region = random.choice(regions)
            salesperson = random.choice(salespeople)
            quantity = random.randint(1, 50)
            unit_price = product[2]
            total = quantity * unit_price

            month = random.randint(1, 12)
            day = random.randint(1, 28)
            date_str = f"2026-{month:02d}-{day:02d}"

            writer.writerow([
                f"ORD-{i+1:04d}",
                product[0],
                product[1],
                region,
                salesperson,
                quantity,
                unit_price,
                total,
                date_str
            ])

    print(f"模拟数据已生成: {csv_path}  (500 条记录)")

    # ─── 2. 报表生成 ───

    print("\n--- 2. 生成报表 ---")

    reporter = ReportGenerator(str(csv_path))
    reporter.clean_data()
    reporter.analyze(group_by="区域", value_fields=["销量", "总金额"])

    # 生成 Excel 报表
    excel_path = tmp_dir / "销售报表.xlsx"
    reporter.to_excel(str(excel_path))

    # 生成 CSV 备份
    backup_path = tmp_dir / "销售数据_清洗后.csv"
    reporter.to_csv(str(backup_path))

    # ─── 3. 使用 pandas 额外分析 ───

    if HAS_PANDAS:
        print("\n--- 3. pandas 补充分析 ---")

        df = pd.read_csv(csv_path, encoding="utf-8-sig")
        df["日期"] = pd.to_datetime(df["日期"])

        print("\n各产品销售总额:")
        top_products = df.groupby("产品名")["总金额"].sum().sort_values(ascending=False)
        for name, total in top_products.head(5).items():
            print(f"  {name:20s}: ¥{total:>10,.2f}")

        print("\n各区域销售额与占比:")
        region_summary = df.groupby("区域").agg(
            销售额=("总金额", "sum"),
            订单数=("订单号", "count")
        ).sort_values("销售额", ascending=False)
        region_summary["占比"] = (region_summary["销售额"] /
                                   region_summary["销售额"].sum() * 100)

        for region, row in region_summary.iterrows():
            print(f"  {region:6s}: ¥{row['销售额']:>10,.2f}  "
                  f"({row['占比']:.1f}%) 订单数: {row['订单数']}")

        print("\n销售员业绩排名:")
        sales_summary = df.groupby("销售员")["总金额"].sum().sort_values(ascending=False)
        for name, total in sales_summary.items():
            print(f"  {name:6s}: ¥{total:>10,.2f}")

    # ─── 清理 ───
    import shutil
    shutil.rmtree(tmp_dir)

    print(f"\n{'=' * 60}")
    print("✅ 报表生成器演示完成！")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
